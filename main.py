import openpyxl
import win32com.client as win32
import os

def save_to_excel(tasks, filename="tasks.xlsx"):
    # Dosyanın tam yolunu belirleyin
    filepath = r"C:\Users\emreg\OneDrive\Masaüstü\deneme\todolist-cli\tasks.xlsx"
    
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Tasks"

    sheet.append(["Task", "Status"])

    for task in tasks:
        status = "Completed" if task["completed"] else "Continue..."
        sheet.append([task["task"], status])

    wb.save(filepath)
    print(f"Tasks saved to {filepath}")

    # Excel dosyasını aç ve yenile
    excel = win32.Dispatch("Excel.Application")
    workbook = excel.Workbooks.Open(filepath)
    excel.CalculateFull()  # Tüm çalışma sayfalarını yenile
    workbook.Save()
    workbook.Close()
    excel.Quit()

def load_from_excel(filename="tasks.xlsx"):
    # Dosyanın tam yolunu belirleyin
    filepath = r"C:\Users\emreg\OneDrive\Masaüstü\deneme\todolist-cli\tasks.xlsx"
    
    tasks = []
    try:
        wb = openpyxl.load_workbook(filepath)
        sheet = wb.active
        
        for row in sheet.iter_rows(min_row=2, values_only=True):
            task, status = row
            tasks.append({"task": task, "completed": status == "Completed"})
        
        print(f"Tasks loaded from {filepath}")
    except FileNotFoundError:
        print(f"{filepath} not found. Starting with an empty task list.")
    
    return tasks

def menu():
    print("\nPICK ONE AND START")
    print("1. Add task")
    print("2. List tasks")
    print("3. Complete task")
    print("4. Delete task")
    print("5. EXIT\n")
    
def add_task(tasks):
    task = input("ADD YOUR TASK: ")
    tasks.append({"task": task, "completed": False})
    save_to_excel(tasks)
    print(f"'{task}' added.")

def list_tasks(tasks):
    if not tasks:
        print("There is no task to do.")
    else:
        for idx, task in enumerate(tasks, 1):
            status = "Completed" if task["completed"] else "Continue..."
            print(f"{idx}. {task['task']} [{status}]")

def delete_task(tasks):
    list_tasks(tasks)
    task_number = int(input("Which number do you want to delete: ")) - 1
    if 0 <= task_number < len(tasks):
        deleted_task = tasks.pop(task_number)
        save_to_excel(tasks)
        print(f"'{deleted_task['task']}' deleted.")
    else:
        print("Invalid task number.")
    
def complete_task(tasks):
    list_tasks(tasks)
    task_number = int(input("Which number do you want to mark as completed: ")) - 1
    if 0 <= task_number < len(tasks):
        tasks[task_number]["completed"] = True
        save_to_excel(tasks)
        print(f"'{tasks[task_number]['task']}' marked as completed.")
    else:
        print("Invalid task number.")
    
def main():
    tasks = load_from_excel()
    while True:
        menu()
        choice = input("Choose one: ")

        match choice:
            case "1":
                add_task(tasks)
            case "2":
                list_tasks(tasks)
            case "3":
                complete_task(tasks)
            case "4":
                delete_task(tasks)
            case "5":
                save_to_excel(tasks)
                print("Exiting...")
                break
            case _:
                print("Invalid choice. Please try again.")

if __name__ == "__main__":
    main()
