import pandas as pd
import openpyxl
import os

class TaskManager:
    def __init__(self, filepath):
        self.filepath = filepath
        self.tasks = self.load_from_excel()

    def save_to_excel(self):
        try:
            # Create a new workbook and select the active sheet
            wb = openpyxl.Workbook()
            sheet = wb.active
            sheet.title = "Tasks"
            sheet.append(["Task", "Status"])

            # Add tasks to the sheet
            for task in self.tasks:
                status = "Completed" if task["completed"] else "Continue..."
                sheet.append([task["task"], status])

            # Save the workbook
            wb.save(self.filepath)
            print(f"Tasks saved to {self.filepath}")
        except Exception as e:
            print(f"Failed to save tasks: {e}")

    def load_from_excel(self):
        tasks = []
        if os.path.exists(self.filepath):
            try:
                wb = openpyxl.load_workbook(self.filepath)
                sheet = wb.active

                # Load tasks from the sheet, skipping the header row
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    task, status = row
                    tasks.append({"task": task, "completed": status == "Completed"})

                print(f"Tasks loaded from {self.filepath}")
            except Exception as e:
                print(f"Failed to load tasks: {e}")
        else:
            print(f"{self.filepath} not found. Starting with an empty task list.")
        return tasks

    def add_task(self, task):
        if task.strip() == "":
            print("Task cannot be empty.")
            return
        self.tasks.append({"task": task, "completed": False})
        self.save_to_excel()
        print(f"'{task}' added.")

    def list_tasks(self):
        if not self.tasks:
            print("There is no task to do.")
        else:
            for idx, task in enumerate(self.tasks, 1):
                status = "Completed" if task["completed"] else "Continue..."
                print(f"{idx}. {task['task']} [{status}]")

    def delete_task(self, task_number):
        if 0 <= task_number < len(self.tasks):
            deleted_task = self.tasks.pop(task_number)
            self.save_to_excel()
            print(f"'{deleted_task['task']}' deleted.")
        else:
            print("Invalid task number.")

    def complete_task(self, task_number):
        if 0 <= task_number < len(self.tasks):
            self.tasks[task_number]["completed"] = True
            self.save_to_excel()
            print(f"'{self.tasks[task_number]['task']}' marked as completed.")
        else:
            print("Invalid task number.")

def menu():
    print("\nPICK ONE AND START")
    print("1. Add task")
    print("2. List tasks")
    print("3. Complete task")
    print("4. Delete task")
    print("5. EXIT\n")

def main():
    filepath = "tasks.xlsx"  # Save in the current directory
    task_manager = TaskManager(filepath)

    while True:
        menu()
        choice = input("Choose one: ")

        if choice == "1":
            task = input("ADD YOUR TASK: ")
            task_manager.add_task(task)
        elif choice == "2":
            task_manager.list_tasks()
        elif choice == "3":
            task_manager.list_tasks()
            try:
                task_number = int(input("Which number do you want to mark as completed: ")) - 1
                task_manager.complete_task(task_number)
            except ValueError:
                print("Invalid input. Please enter a number.")
        elif choice == "4":
            task_manager.list_tasks()
            try:
                task_number = int(input("Which number do you want to delete: ")) - 1
                task_manager.delete_task(task_number)
            except ValueError:
                print("Invalid input. Please enter a number.")
        elif choice == "5":
            task_manager.save_to_excel()
            print("Exiting...")
            break
        else:
            print("Invalid choice. Please try again.")

if __name__ == "__main__":
    main()
